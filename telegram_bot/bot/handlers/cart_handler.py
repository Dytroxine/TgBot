import os
import asyncio
from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery, Message
from utils.db_helpers import get_db_connection
from utils.logger import logger
import stripe
from aiogram.fsm.context import FSMContext
from utils.subscription_checker import require_subscription
from openpyxl import load_workbook, Workbook


stripe.api_key = os.getenv("STRIPE_API_KEY")
user_waiting_for_data = {}
dp = Dispatcher()
@require_subscription
async def add_to_cart(callback_query: CallbackQuery):
    """Добавление товара в корзину."""
    product_id = int(callback_query.data.split("_")[2])
    user_id = callback_query.from_user.id

    connection = await get_db_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        INSERT INTO cart (user_id, product_id, quantity)
        VALUES (%s, %s, 1)
        ON CONFLICT (user_id, product_id)
        DO UPDATE SET quantity = cart.quantity + 1;
        """,
        (user_id, product_id)
    )
    connection.commit()
    connection.close()

    await callback_query.answer("Товар добавлен в корзину! 🛒")


@require_subscription
async def view_cart(message: Message, is_edit: bool = False, cart_items=None):
    """Просмотр содержимого корзины с возможностью обновления."""
    user_id = message.from_user.id

    # Если cart_items не переданы, загружаем из базы
    if cart_items is None:
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(
            """
            SELECT p.id, p.name, c.quantity, p.price, (p.price * c.quantity) AS total
            FROM cart c
            JOIN products p ON c.product_id = p.id
            WHERE c.user_id = %s;
            """,
            (user_id,)
        )
        cart_items = cursor.fetchall()
        connection.close()

    # Если корзина пуста, обновляем сообщение
    if not cart_items:
        cart_text = "**Ваша корзина:**\n\nОбщая сумма: 0 ₽"
        keyboard = [[InlineKeyboardButton(text="Оформить заказ 🛒", callback_data="cart_order")]]
        reply_markup = InlineKeyboardMarkup(inline_keyboard=keyboard)

        if is_edit:
            try:
                await message.edit_text(text=cart_text, parse_mode="Markdown", reply_markup=reply_markup)
            except Exception as e:
                print(f"Ошибка редактирования сообщения: {e}")  # Логирование ошибки
        else:
            await message.answer(text=cart_text, parse_mode="Markdown", reply_markup=reply_markup)
        return

    # Строим сообщение с товарами
    cart_text = "**Ваша корзина:**\n\n"
    total_sum = 0
    keyboard = []

    for item in cart_items:
        product_id, name, quantity, price, total = item
        cart_text += (
            f"{name}\n"
            f"Цена: {price} ₽\n"
            f"Количество: {quantity}\n"
            f"Итого: {total} ₽\n\n"
        )
        total_sum += total

        # Кнопки управления для каждого товара
        keyboard.append([
            InlineKeyboardButton(text="➖", callback_data=f"cart_decrease_{product_id}"),
            InlineKeyboardButton(text=str(quantity), callback_data=f"noop_{product_id}"),
            InlineKeyboardButton(text="➕", callback_data=f"cart_increase_{product_id}")
        ])

    # Кнопка "Оформить заказ"
    keyboard.append([InlineKeyboardButton(text="Оформить заказ 🛒", callback_data=f"cart_order_{total_sum}")])

    cart_text += f"**Общая сумма: {total_sum} ₽**"
    reply_markup = InlineKeyboardMarkup(inline_keyboard=keyboard)

    if is_edit:
        try:
            await message.edit_text(text=cart_text, parse_mode="Markdown", reply_markup=reply_markup)
        except Exception as e:
            print(f"Ошибка редактирования сообщения: {e}")  # Логирование ошибки
    else:
        await message.answer(text=cart_text, parse_mode="Markdown", reply_markup=reply_markup)








def write_to_excel(user_id, delivery_name, delivery_address, delivery_phone, total_sum):
    file_name = "orders.xlsx"
    logger.info("Запись данных в Excel")
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["User ID", "Имя", "Адрес доставки", "Телефон", "Сумма заказа"])

    sheet.append([user_id, delivery_name, delivery_address, delivery_phone, total_sum])
    workbook.save(file_name)
    workbook.close()


async def handle_delivery_data(callback_query: CallbackQuery, total_sum):
    user_id = callback_query.from_user.id

    # Проверяем, есть ли данные для записи
    if total_sum == 0:
        await callback_query.message.answer("Ошибка: Корзина пуста или сумма заказа неизвестна.")
        return

    # Сообщаем пользователю, что ждем данные
    await callback_query.message.answer(
        "Пожалуйста, введите все данные в формате:\n"
        "1. Ваше имя и фамилия\n"
        "2. Адрес доставки\n"
        "3. Ваш номер телефона"
    )
    user_waiting_for_data[user_id] = total_sum


async def process_user_input(message: Message):
    user_id = message.from_user.id

    # Проверяем, ждем ли данные от этого пользователя
    if user_id in user_waiting_for_data:
        delivery_data = message.text.split("\n")
        if len(delivery_data) < 3:
            await message.answer(
                "Некорректный формат. Пожалуйста, отправьте данные в формате:\n"
                "1. Ваше имя и фамилия\n"
                "2. Адрес доставки\n"
                "3. Ваш номер телефона"
            )
            return

        # Получаем сумму заказа
        total_sum = user_waiting_for_data.pop(user_id)

        # Записываем данные в Excel (замените write_to_excel на свою функцию)
        write_to_excel(user_id, delivery_data[0], delivery_data[1], delivery_data[2], total_sum)

        # Подтверждаем успешное получение данных
        await message.answer("Спасибо! Теперь вы будете перенаправлены на платёж.")

        # Инициация платежа (замените initiate_payment на свою функцию)
        await initiate_payment(message)



async def initiate_payment(callback_query: CallbackQuery):
    """Инициация платежа через Stripe."""
    user_id = callback_query.from_user.id
    connection = get_db_connection()
    cursor = connection.cursor()

    cursor.execute(
        """
        SELECT p.name, c.quantity, p.price, (p.price * c.quantity) AS total
        FROM cart c
        JOIN products p ON c.product_id = p.id
        WHERE c.user_id = %s;
        """,
        (user_id,)
    )
    cart_items = cursor.fetchall()
    connection.close()

    if not cart_items:
        await callback_query.answer("Ваша корзина пуста.")
        return

    total_sum = sum(item[3] for item in cart_items)

    try:
        session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'rub',
                    'product_data': {
                        'name': 'Оплата заказа'
                    },
                    'unit_amount': int(total_sum * 100),
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url='https://example.com/success',
            cancel_url='https://example.com/cancel',
        )
        await callback_query.answer(
            "Нажмите на кнопку ниже, чтобы перейти к оплате:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="Оплатить", url=session.url)
            ]])
        )
    except Exception as e:
        logger.error(f"Ошибка при создании платежной сессии: {e}")
        await callback_query.answer("Произошла ошибка при создании платежа.")


async def handle_cart_callback(callback_query: CallbackQuery):
    """Обработка callback'ов корзины."""
    await callback_query.answer()
    user_id = callback_query.from_user.id
    data = callback_query.data.split("_")

    action = data[1]  # "increase" или "decrease"
    if action == "order":
        await handle_delivery_data(callback_query, data[2])
        return
    product_id = int(data[2])  # Уникальный идентификатор товара

    connection = get_db_connection()
    cursor = connection.cursor()

        # Получаем количество текущего товара
    cursor.execute("""
        SELECT quantity 
        FROM cart
        WHERE user_id = %s AND product_id = %s;
    """, (user_id, product_id))
    product_row = cursor.fetchone()

    if not product_row:
        await callback_query.message.answer("Товар не найден в корзине.")
        connection.close()
        return

    current_quantity = product_row[0]

    # Обновляем количество товара в базе
    if action == "increase":
        cursor.execute("""
            UPDATE cart
            SET quantity = quantity + 1
            WHERE user_id = %s AND product_id = %s;
        """, (user_id, product_id))
    elif action == "decrease":
        if current_quantity > 1:
            cursor.execute("""
                UPDATE cart
                SET quantity = quantity - 1
                WHERE user_id = %s AND product_id = %s;
            """, (user_id, product_id))
        else:
            cursor.execute("""
                DELETE FROM cart
                WHERE user_id = %s AND product_id = %s;
            """, (user_id, product_id))

    # Фиксируем изменения в базе
    connection.commit()

    # **Повторно загружаем данные после обновления**
    cursor.execute("""
        SELECT p.id, p.name, c.quantity, p.price, (p.price * c.quantity) AS total
        FROM cart c
        JOIN products p ON c.product_id = p.id
        WHERE c.user_id = %s;
    """, (user_id,))
    updated_cart_items = cursor.fetchall()
    connection.close()

    # **Вызываем обновление корзины, передавая обновлённые данные**
    await view_cart(callback_query.message, is_edit=True, cart_items=updated_cart_items)



